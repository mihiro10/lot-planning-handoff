"""
Shared handoff logic used by streamlit_app.py, webapp.py, and handoff.py.
All business logic lives here; entry-point files contain only UI/IO.
"""

from io import BytesIO
from datetime import datetime, timedelta, date
import traceback
import openpyxl
from openpyxl.cell.cell import MergedCell

COL_CODE        = 5
COL_NAME        = 11  # 品目名
COL_H           = 7   # 棚卸し前在庫
COL_L           = 12  # ロット（一回あたり）
COL_R           = 18  # 入庫リードタイム（日）
COL_RTYPE       = 20  # 行種別
COL_DAY1        = 21
MAX_LEAD_TIME   = 6
TRANSFER_TYPES  = ['備考', '計画（倍）', '使用予測']


def parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    return None


def safe_write(ws, row, col, val):
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        cell.value = val


def find_planning_sheet(wb):
    """Return the sheet with the latest date in U2; falls back to wb.active."""
    best_ws   = None
    best_date = None
    for ws in wb.worksheets:
        d = parse_date(ws.cell(2, COL_DAY1).value)
        if d and (best_date is None or d >= best_date):
            best_ws   = ws
            best_date = d
    return best_ws or wb.active


def build_product_map(ws):
    products = {}
    current_code = None
    for row in range(3, ws.max_row + 1):
        rtype = ws.cell(row, COL_RTYPE).value
        code  = ws.cell(row, COL_CODE).value
        if not rtype:
            continue
        if code:
            current_code = code
        if not current_code:
            continue
        if current_code not in products:
            products[current_code] = {'name': None, 'blocks': [{}]}
        p = products[current_code]
        if rtype == '備考' and p['blocks'][-1]:
            p['blocks'].append({})
        p['blocks'][-1][rtype] = row
        if rtype == '備考':
            p['blocks'][-1]['_lot_size']  = ws.cell(row, COL_L).value or 0
            p['blocks'][-1]['_lead_time'] = int(ws.cell(row, COL_R).value or 0)
            if code:
                p['name'] = ws.cell(row, COL_NAME).value
    return products


def run_handoff(may_bytes, jun_bytes, may_sheet=None, jun_sheet=None):
    """
    Transfer overflow planning data from 今月 to 来月.

    Returns a result dict with keys:
      success, error, detail,
      may_start, jun_start, may_last_date,
      transferred, new_products, discontinued, warnings,
      jun_bytes
    """
    result = {
        'success': False, 'error': None, 'detail': None,
        'may_start': None, 'jun_start': None, 'may_last_date': None,
        'transferred': [], 'new_products': [], 'discontinued': [], 'warnings': [],
        'jun_bytes': None,
    }
    try:
        may_wb = openpyxl.load_workbook(BytesIO(may_bytes), data_only=True)
        jun_wb = openpyxl.load_workbook(BytesIO(jun_bytes))
        may_ws = may_wb[may_sheet] if may_sheet else find_planning_sheet(may_wb)
        jun_ws = jun_wb[jun_sheet] if jun_sheet else find_planning_sheet(jun_wb)

        may_start = parse_date(may_ws.cell(2, COL_DAY1).value)
        jun_start = parse_date(jun_ws.cell(2, COL_DAY1).value)

        if not may_start:
            result['error'] = f'今月ファイルのシート「{may_ws.title}」のU2セルに日付が見つかりません。日付形式を確認してください。'
            return result
        if not jun_start:
            result['error'] = f'来月ファイルのシート「{jun_ws.title}」のU2セルに日付が見つかりません。日付形式を確認してください。'
            return result
        if jun_start <= may_start:
            result['error'] = f'来月の開始日（{jun_start}）が今月の開始日（{may_start}）より前です。ファイルの順番を確認してください。'
            return result

        result['may_start']     = str(may_start)
        result['jun_start']     = str(jun_start)
        may_last_date           = jun_start - timedelta(days=1)
        result['may_last_date'] = str(may_last_date)

        is_last_day       = date.today() >= may_last_date
        may_last_col      = COL_DAY1 + (may_last_date - may_start).days
        overlap_start_col = COL_DAY1 + (jun_start - may_start).days

        if overlap_start_col > may_ws.max_column:
            result['error'] = f'今月ファイルにオーバーフロー列（{jun_start}以降）が見つかりません。今月ファイルに翌月分の列が含まれているか確認してください。'
            return result

        overlap_days = min(
            may_ws.max_column - overlap_start_col + 1,
            jun_ws.max_column - COL_DAY1 + 1,
        )

        may_products = build_product_map(may_ws)
        jun_products = build_product_map(jun_ws)

        if not may_products:
            result['error'] = '今月ファイルにコード（E列）が入力された商品が見つかりません。'
            return result
        if not jun_products:
            result['error'] = '来月ファイルにコード（E列）が入力された商品が見つかりません。'
            return result

        for code, info in may_products.items():
            if code not in jun_products:
                result['discontinued'].append({'code': code, 'name': info.get('name') or '（名前なし）'})

        for code, jun_info in jun_products.items():
            name = jun_info['name'] or '（名前なし）'
            if code not in may_products:
                result['new_products'].append({'code': code, 'name': name})
                continue

            may_blocks = may_products[code]['blocks']
            jun_blocks = jun_info['blocks']
            item = {
                'code': code, 'name': name,
                'takadoshi_mae': None, 'takadoshi_warning': False,
                'transferred_types': [], 'skipped_types': [],
            }
            nyuko_written_dates = []
            product_any_val = False

            for b_idx in range(min(len(may_blocks), len(jun_blocks))):
                may_block = may_blocks[b_idx]
                jun_block = jun_blocks[b_idx]

                if is_last_day:
                    if '最終' not in may_block:
                        result['warnings'].append(f'[{code}] {name} ブロック{b_idx+1}: 今月に最終行が見つかりません。棚卸し前在庫をスキップしました。')
                    elif '備考' not in jun_block:
                        result['warnings'].append(f'[{code}] {name} ブロック{b_idx+1}: 来月に備考行が見つかりません。棚卸し前在庫をスキップしました。')
                    else:
                        raw = may_ws.cell(may_block['最終'], may_last_col).value
                        if isinstance(raw, str) and raw.startswith('#'):
                            pass  # formula error in source — skip silently
                        else:
                            safe_write(jun_ws, jun_block['備考'], COL_H, raw)
                            if b_idx == 0:
                                item['takadoshi_mae'] = raw
                            if raw is None:
                                item['takadoshi_warning'] = True
                                result['warnings'].append(f'[{code}] {name} ブロック{b_idx+1}: 今月末（{may_last_date}）の最終在庫が空白です。手動で確認してください。')

                for rtype in TRANSFER_TYPES:
                    if rtype not in may_block:
                        if b_idx == 0:
                            item['skipped_types'].append(f'{rtype}（今月に行なし）')
                        continue
                    if rtype not in jun_block:
                        if b_idx == 0:
                            item['skipped_types'].append(f'{rtype}（来月に行なし）')
                        continue
                    any_val = False
                    for i in range(overlap_days):
                        val = may_ws.cell(may_block[rtype], overlap_start_col + i).value
                        safe_write(jun_ws, jun_block[rtype], COL_DAY1 + i, val)
                        if val is not None:
                            any_val = True
                            product_any_val = True
                    if any_val and rtype not in item['transferred_types']:
                        item['transferred_types'].append(rtype)

                lead_time = min(may_block.get('_lead_time', 0), MAX_LEAD_TIME)
                lot_size  = may_block.get('_lot_size', 0)
                if lead_time and lot_size and '入庫予定数' in jun_block and '計画（倍）' in may_block:
                    for d in range(lead_time):
                        may_col = overlap_start_col - lead_time + d
                        if may_col < COL_DAY1:
                            continue
                        keikaku = may_ws.cell(may_block['計画（倍）'], may_col).value
                        if keikaku:
                            safe_write(jun_ws, jun_block['入庫予定数'], COL_DAY1 + d, keikaku * lot_size)
                            if b_idx == 0:
                                nyuko_written_dates.append(jun_start + timedelta(days=d))

            if not product_any_val:
                result['warnings'].append(f'[{code}] {name}: 今月のオーバーフロー欄にデータがありませんでした。')

            if nyuko_written_dates:
                first = nyuko_written_dates[0]
                last  = nyuko_written_dates[-1]
                if first == last:
                    label = f'入庫予定数（{first.month}/{first.day}）'
                else:
                    label = f'入庫予定数（{first.month}/{first.day}〜{last.month}/{last.day}）'
                item['transferred_types'].append(label)

            result['transferred'].append(item)

        jun_ws.protection.sheet = False

        out = BytesIO()
        jun_wb.save(out)
        out.seek(0)
        result['jun_bytes'] = out.getvalue()
        result['success'] = True

    except Exception as e:
        result['error'] = f'予期しないエラーが発生しました: {e}'
        result['detail'] = traceback.format_exc()

    return result
