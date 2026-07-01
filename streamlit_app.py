import streamlit as st
from io import BytesIO
import openpyxl
from core import run_handoff, find_planning_sheet

# ── UI ──────────────────────────────────────────────────────────────────────

st.set_page_config(page_title='月次引き継ぎ', page_icon='📋', layout='centered')
st.title('月次引き継ぎ')
st.caption('今月のオーバーフロー計画（備考・計画・入庫・使用予測）と今月末在庫を来月ファイルに転記します。')

with st.expander('このアプリの使い方', expanded=False):
    st.markdown("""
**月末に今月のデータを来月ファイルへ引き継ぐツールです。**

#### 転記される内容
| 項目 | 説明 |
|------|------|
| 備考 | 今月ファイルの来月分オーバーフロー欄の備考を転記 |
| 計画（倍） | 今月ファイルの来月分オーバーフロー計画を転記 |
| 使用予測 | 今月ファイルの来月分オーバーフロー使用予測を転記 |
| 入庫予定数 | 入庫リードタイム（R列）分だけ来月初日の入庫予定を補正（最大6日） |
| 棚卸し前在庫 | 月末最終日のみ・今月最終在庫を来月ファイルの棚卸し前在庫欄に転記 |

#### 手順
1. **今月のファイル（記入済み）** と **来月のファイル（空白テンプレート）** をアップロード
2. 複数シートがある場合はドロップダウンで対象シートを選択（自動検出）
3. 「引き継ぎを実行」をクリック
4. 結果を確認し、**来月のファイルをダウンロード（記入済み）** で保存

#### 注意事項
- 来月ファイルのシート保護は自動的に解除されます
- ダウンロードされるファイル名は元のファイル名に「移行済」が付きます
- 新規商品・廃止商品は自動では処理されないため、手動対応が必要です
""")

col1, col2 = st.columns(2)
with col1:
    may_file = st.file_uploader('今月のファイル（記入済み）（例: 6月計画.xlsx）', type='xlsx', key='may')
with col2:
    jun_file = st.file_uploader('来月のファイル（空白）（例: 7月計画.xlsx）', type='xlsx', key='jun')

may_sheet = None
jun_sheet = None

if may_file and jun_file:
    may_bytes = may_file.read()
    jun_bytes = jun_file.read()

    may_wb_peek = openpyxl.load_workbook(BytesIO(may_bytes), read_only=True, data_only=True)
    jun_wb_peek = openpyxl.load_workbook(BytesIO(jun_bytes), read_only=True, data_only=True)

    may_sheets = may_wb_peek.sheetnames
    jun_sheets = jun_wb_peek.sheetnames

    may_default = find_planning_sheet(openpyxl.load_workbook(BytesIO(may_bytes), data_only=True)).title
    jun_default = find_planning_sheet(openpyxl.load_workbook(BytesIO(jun_bytes), data_only=True)).title

    col3, col4 = st.columns(2)
    with col3:
        may_sheet = st.selectbox('今月：使用するシート', may_sheets, index=may_sheets.index(may_default))
    with col4:
        jun_sheet = st.selectbox('来月：使用するシート', jun_sheets, index=jun_sheets.index(jun_default))

if st.button('引き継ぎを実行', type='primary', disabled=not (may_file and jun_file)):
    with st.spinner('処理中...'):
        result = run_handoff(may_bytes, jun_bytes, may_sheet, jun_sheet)

    if result['error']:
        st.error(result['error'])
        if result.get('detail'):
            with st.expander('詳細エラー'):
                st.code(result['detail'])
        st.stop()

    # Summary bar
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric('今月末', result['may_last_date'])
    c2.metric('来月開始', result['jun_start'])
    c3.metric('転記済み', len(result['transferred']))
    c4.metric('新規', len(result['new_products']))
    c5.metric('廃止', len(result['discontinued']))

    # Download
    stem = jun_file.name.rsplit('.', 1)[0]
    st.download_button(
        '📥 来月のファイルをダウンロード（記入済み）',
        data=result['jun_bytes'],
        file_name=f'{stem}移行済.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        type='primary',
    )

    # Duplicate codes
    for label, dupes in [('今月', result['may_duplicates']), ('来月', result['jun_duplicates'])]:
        if dupes:
            st.error(f'{label}ファイルに重複コードがあります。確認してください。')
            st.dataframe(
                [{'コード': d['code'], '該当品目名': d['names']} for d in dupes],
                use_container_width=True, hide_index=True,
            )

    # Unified results table
    if result['transferred']:
        st.subheader('転記結果')
        rows = []
        for item in result['transferred']:
            transferred = '、'.join(item['transferred_types']) if item['transferred_types'] else '（なし）'
            takadoshi = item['takadoshi_mae'] if item['takadoshi_mae'] is not None else '―'
            notes = '　'.join(item['notes']) if item['notes'] else ''
            rows.append({
                'コード': str(item['code']),
                '品目名': item['name'],
                '転記した行': transferred,
                '前月最終在庫・次月棚卸し前在庫': str(takadoshi),
                '備考': notes,
            })
        st.dataframe(rows, use_container_width=True, hide_index=True)

    if result['new_products']:
        st.subheader(f'新規商品（{len(result["new_products"])}件） — 在庫数を手動入力してください')
        st.dataframe(
            [{'コード': str(p['code']), '品目名': p['name']} for p in result['new_products']],
            use_container_width=True, hide_index=True,
        )

    if result['discontinued']:
        st.subheader(f'廃止商品（{len(result["discontinued"])}件） — 来月ファイルに存在しないためスキップ')
        st.dataframe(
            [{'コード': str(p['code']), '品目名': p['name']} for p in result['discontinued']],
            use_container_width=True, hide_index=True,
        )
