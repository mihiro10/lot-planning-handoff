"""
月次引き継ぎスクリプト
使い方: python3 handoff.py <今月ファイル.xlsx> <来月ファイル.xlsx>

今月のオーバーフロー計画（備考・計画（倍）・使用予測）を来月ファイルに転記し、
月末の場合のみ今月末の最終在庫を来月の棚卸し前在庫にセットします。
"""

import sys
import openpyxl
from openpyxl.cell.cell import MergedCell
from datetime import datetime, timedelta, date

COL_CODE        = 5
COL_NAME        = 11  # 品目名
COL_H           = 7   # 棚卸し前在庫
COL_L           = 12  # ロット（一回あたり）
COL_R           = 18  # 入庫リードタイム（日）
COL_RTYPE       = 20  # 行種別
COL_DAY1        = 21
MAX_LEAD_TIME   = 6
TRANSFER_TYPES  = ['備考', '計画（倍）', '使用予測']


def parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    return None


def safe_write(ws, row, col, val):
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        cell.value = val


def find_planning_sheet(wb):
    for ws in wb.worksheets:
        if isinstance(ws.cell(2, COL_DAY1).value, datetime):
            return ws
    return wb.active


def build_product_map(ws):
    products = {}
    current_code = None
    for row in range(3, ws.max_row + 1):
        rtype = ws.cell(row, COL_RTYPE).value
        code  = ws.cell(row, COL_CODE).value
        if not rtype:
            continue
        if code:
            current_code = code
        if not current_code:
            continue
        if current_code not in products:
            products[current_code] = {'name': None, 'blocks': [{}]}
        p = products[current_code]
        if rtype == '備考' and p['blocks'][-1]:
            p['blocks'].append({})
        p['blocks'][-1][rtype] = row
        if rtype == '備考':
            p['blocks'][-1]['_lot_size']  = ws.cell(row, COL_L).value or 0
            p['blocks'][-1]['_lead_time'] = int(ws.cell(row, COL_R).value or 0)
            if code:
                p['name'] = ws.cell(row, COL_NAME).value
    return products


def run_handoff(may_path, jun_path):
    print(f"今月: {may_path}")
    print(f"来月: {jun_path}")

    may_wb = openpyxl.load_workbook(may_path, data_only=True)
    jun_wb = openpyxl.load_workbook(jun_path)
    may_ws = find_planning_sheet(may_wb)
    jun_ws = find_planning_sheet(jun_wb)

    print(f"今月シート: {may_ws.title}")
    print(f"来月シート: {jun_ws.title}")

    may_start = parse_date(may_ws.cell(2, COL_DAY1).value)
    jun_start = parse_date(jun_ws.cell(2, COL_DAY1).value)

    if not may_start:
        print(f"ERROR: 今月ファイルのシート「{may_ws.title}」のU2セルに日付が見つかりません。")
        sys.exit(1)
    if not jun_start:
        print(f"ERROR: 来月ファイルのシート「{jun_ws.title}」のU2セルに日付が見つかりません。")
        sys.exit(1)
    if jun_start <= may_start:
        print(f"ERROR: 来月の開始日({jun_start})が今月の開始日({may_start})より前です。ファイルの順番を確認してください。")
        sys.exit(1)

    may_last_date     = jun_start - timedelta(days=1)
    is_last_day       = date.today() == may_last_date
    may_last_col      = COL_DAY1 + (may_last_date - may_start).days
    overlap_start_col = COL_DAY1 + (jun_start - may_start).days

    if overlap_start_col > may_ws.max_column:
        print(f"ERROR: 今月ファイルにオーバーフロー列（{jun_start}以降）が見つかりません。")
        sys.exit(1)

    overlap_days = min(
        may_ws.max_column - overlap_start_col + 1,
        jun_ws.max_column - COL_DAY1 + 1,
    )

    print(f"今月開始: {may_start}  今月末: {may_last_date}  来月開始: {jun_start}")
    print(f"オーバーフロー開始: col {overlap_start_col}  期間: {overlap_days}日  本日が月末: {is_last_day}")

    may_products = build_product_map(may_ws)
    jun_products = build_product_map(jun_ws)

    if not may_products:
        print("ERROR: 今月ファイルにコード（E列）が入力された商品が見つかりません。")
        sys.exit(1)
    if not jun_products:
        print("ERROR: 来月ファイルにコード（E列）が入力された商品が見つかりません。")
        sys.exit(1)

    warnings = []

    for code, info in may_products.items():
        if code not in jun_products:
            print(f"  廃止: [{code}] {info.get('name') or ''}")

    for code, jun_info in jun_products.items():
        name = jun_info['name'] or '（名前なし）'
        if code not in may_products:
            print(f"  新規（手動入力が必要）: [{code}] {name}")
            continue

        may_blocks = may_products[code]['blocks']
        jun_blocks = jun_info['blocks']
        nyuko_written_dates = []

        for b_idx in range(min(len(may_blocks), len(jun_blocks))):
            may_block = may_blocks[b_idx]
            jun_block = jun_blocks[b_idx]

            if is_last_day:
                if '最終' not in may_block:
                    warnings.append(f'[{code}] {name} ブロック{b_idx+1}: 最終行が見つかりません。棚卸し前在庫をスキップしました。')
                elif '備考' not in jun_block:
                    warnings.append(f'[{code}] {name} ブロック{b_idx+1}: 来月に備考行が見つかりません。棚卸し前在庫をスキップしました。')
                else:
                    val = may_ws.cell(may_block['最終'], may_last_col).value
                    safe_write(jun_ws, jun_block['備考'], COL_H, val)
                    if b_idx == 0:
                        print(f"  [{code}] 棚卸し前在庫 = {val}")
                    if val is None:
                        warnings.append(f'[{code}] {name} ブロック{b_idx+1}: 今月末({may_last_date})の最終在庫が空白です。手動で確認してください。')

            block_any_val = False
            for rtype in TRANSFER_TYPES:
                if rtype not in may_block or rtype not in jun_block:
                    continue
                for i in range(overlap_days):
                    val = may_ws.cell(may_block[rtype], overlap_start_col + i).value
                    safe_write(jun_ws, jun_block[rtype], COL_DAY1 + i, val)
                    if val is not None:
                        block_any_val = True

            if b_idx == 0 and not block_any_val:
                warnings.append(f'[{code}] {name}: 今月のオーバーフロー欄にデータがありませんでした。')

            lead_time = min(may_block.get('_lead_time', 0), MAX_LEAD_TIME)
            lot_size  = may_block.get('_lot_size', 0)
            if lead_time and lot_size and '入庫予定数' in jun_block and '計画（倍）' in may_block:
                for d in range(lead_time):
                    may_col = overlap_start_col - lead_time + d
                    if may_col < COL_DAY1:
                        continue
                    keikaku = may_ws.cell(may_block['計画（倍）'], may_col).value
                    if keikaku:
                        safe_write(jun_ws, jun_block['入庫予定数'], COL_DAY1 + d, keikaku * lot_size)
                        if b_idx == 0:
                            nyuko_written_dates.append(jun_start + timedelta(days=d))

        transferred_label = code
        if nyuko_written_dates:
            first, last = nyuko_written_dates[0], nyuko_written_dates[-1]
            if first == last:
                transferred_label += f' +入庫予定数補正({first.month}/{first.day})'
            else:
                transferred_label += f' +入庫予定数補正({first.month}/{first.day}〜{last.month}/{last.day})'
        print(f"  転記済み: [{transferred_label}] {name}")

    jun_ws.protection.sheet = False

    stem = jun_path.rsplit('.', 1)[0]
    out_path = f'{stem}移行済.xlsx'
    jun_wb.save(out_path)

    if warnings:
        print(f"\n警告 ({len(warnings)}件):")
        for w in warnings:
            print(f"  ⚠ {w}")

    print(f"\n出力ファイル: {out_path}")


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("使い方: python3 handoff.py <今月.xlsx> <来月.xlsx>")
        sys.exit(1)
    run_handoff(sys.argv[1], sys.argv[2])
