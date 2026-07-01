from flask import Flask, request, send_file, render_template_string
from io import BytesIO
from datetime import datetime, timedelta, date
import traceback
import openpyxl
from openpyxl.cell.cell import MergedCell

app = Flask(__name__)

COL_CODE        = 5
COL_NAME        = 11  # 品目名
COL_H           = 7   # 棚卸し前在庫
COL_L           = 12  # ロット（一回あたり）
COL_R           = 18  # 入庫リードタイム（日）
COL_RTYPE       = 20  # 行種別
COL_DAY1        = 21
MAX_LEAD_TIME   = 6
TRANSFER_TYPES  = ['備考', '計画（倍）', '使用予測']


def parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    return None


def safe_write(ws, row, col, val):
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        cell.value = val


def find_planning_sheet(wb):
    for ws in wb.worksheets:
        if isinstance(ws.cell(2, COL_DAY1).value, datetime):
            return ws
    return wb.active


def build_product_map(ws):
    products = {}
    current_code = None
    for row in range(3, ws.max_row + 1):
        rtype = ws.cell(row, COL_RTYPE).value
        code  = ws.cell(row, COL_CODE).value
        if not rtype:
            continue
        if code:
            current_code = code
        if not current_code:
            continue
        if current_code not in products:
            products[current_code] = {'name': None, 'blocks': [{}]}
        p = products[current_code]
        if rtype == '備考' and p['blocks'][-1]:
            p['blocks'].append({})
        p['blocks'][-1][rtype] = row
        if rtype == '備考':
            p['blocks'][-1]['_lot_size']  = ws.cell(row, COL_L).value or 0
            p['blocks'][-1]['_lead_time'] = int(ws.cell(row, COL_R).value or 0)
            if code:
                p['name'] = ws.cell(row, COL_NAME).value
    return products


def run_handoff(may_bytes, jun_bytes, may_sheet=None, jun_sheet=None):
    result = {
        'success': False, 'error': None, 'detail': None,
        'may_start': None, 'jun_start': None, 'may_last_date': None,
        'transferred': [], 'new_products': [], 'discontinued': [], 'warnings': [],
        'jun_bytes': None,
    }
    try:
        may_wb = openpyxl.load_workbook(BytesIO(may_bytes), data_only=True)
        jun_wb = openpyxl.load_workbook(BytesIO(jun_bytes))
        may_ws = may_wb[may_sheet] if may_sheet else find_planning_sheet(may_wb)
        jun_ws = jun_wb[jun_sheet] if jun_sheet else find_planning_sheet(jun_wb)

        may_start = parse_date(may_ws.cell(2, COL_DAY1).value)
        jun_start = parse_date(jun_ws.cell(2, COL_DAY1).value)

        if not may_start:
            result['error'] = f'今月ファイルのシート「{may_ws.title}」のU2セルに日付が見つかりません。日付形式を確認してください。'
            return result
        if not jun_start:
            result['error'] = f'来月ファイルのシート「{jun_ws.title}」のU2セルに日付が見つかりません。日付形式を確認してください。'
            return result
        if jun_start <= may_start:
            result['error'] = f'来月の開始日（{jun_start}）が今月の開始日（{may_start}）より前です。ファイルの順番を確認してください。'
            return result

        result['may_start']     = str(may_start)
        result['jun_start']     = str(jun_start)
        may_last_date           = jun_start - timedelta(days=1)
        result['may_last_date'] = str(may_last_date)

        is_last_day       = date.today() == may_last_date
        may_last_col      = COL_DAY1 + (may_last_date - may_start).days
        overlap_start_col = COL_DAY1 + (jun_start - may_start).days

        if overlap_start_col > may_ws.max_column:
            result['error'] = f'今月ファイルにオーバーフロー列（{jun_start}以降）が見つかりません。今月ファイルに翌月分の列が含まれているか確認してください。'
            return result

        overlap_days = min(
            may_ws.max_column - overlap_start_col + 1,
            jun_ws.max_column - COL_DAY1 + 1,
        )

        may_products = build_product_map(may_ws)
        jun_products = build_product_map(jun_ws)

        if not may_products:
            result['error'] = '今月ファイルにコード（E列）が入力された商品が見つかりません。'
            return result
        if not jun_products:
            result['error'] = '来月ファイルにコード（E列）が入力された商品が見つかりません。'
            return result

        for code, info in may_products.items():
            if code not in jun_products:
                result['discontinued'].append({'code': code, 'name': info.get('name') or '（名前なし）'})

        for code, jun_info in jun_products.items():
            name = jun_info['name'] or '（名前なし）'
            if code not in may_products:
                result['new_products'].append({'code': code, 'name': name})
                continue

            may_blocks = may_products[code]['blocks']
            jun_blocks = jun_info['blocks']
            item = {
                'code': code, 'name': name,
                'takadoshi_mae': None, 'takadoshi_warning': False,
                'transferred_types': [], 'skipped_types': [],
            }
            nyuko_written_dates = []

            for b_idx in range(min(len(may_blocks), len(jun_blocks))):
                may_block = may_blocks[b_idx]
                jun_block = jun_blocks[b_idx]

                if is_last_day:
                    if '最終' not in may_block:
                        result['warnings'].append(f'[{code}] {name} ブロック{b_idx+1}: 今月に最終行が見つかりません。棚卸し前在庫をスキップしました。')
                    elif '備考' not in jun_block:
                        result['warnings'].append(f'[{code}] {name} ブロック{b_idx+1}: 来月に備考行が見つかりません。棚卸し前在庫をスキップしました。')
                    else:
                        val = may_ws.cell(may_block['最終'], may_last_col).value
                        safe_write(jun_ws, jun_block['備考'], COL_H, val)
                        if b_idx == 0:
                            item['takadoshi_mae'] = val
                        if val is None:
                            item['takadoshi_warning'] = True
                            result['warnings'].append(f'[{code}] {name} ブロック{b_idx+1}: 今月末（{may_last_date}）の最終在庫が空白です。手動で確認してください。')

                block_any_val = False
                for rtype in TRANSFER_TYPES:
                    if rtype not in may_block:
                        if b_idx == 0:
                            item['skipped_types'].append(f'{rtype}（今月に行なし）')
                        continue
                    if rtype not in jun_block:
                        if b_idx == 0:
                            item['skipped_types'].append(f'{rtype}（来月に行なし）')
                        continue
                    any_val = False
                    for i in range(overlap_days):
                        val = may_ws.cell(may_block[rtype], overlap_start_col + i).value
                        safe_write(jun_ws, jun_block[rtype], COL_DAY1 + i, val)
                        if val is not None:
                            any_val = True
                            block_any_val = True
                    if b_idx == 0 and any_val:
                        item['transferred_types'].append(rtype)

                if b_idx == 0 and not block_any_val:
                    result['warnings'].append(f'[{code}] {name}: 今月のオーバーフロー欄にデータがありませんでした。')

                lead_time = min(may_block.get('_lead_time', 0), MAX_LEAD_TIME)
                lot_size  = may_block.get('_lot_size', 0)
                if lead_time and lot_size and '入庫予定数' in jun_block and '計画（倍）' in may_block:
                    for d in range(lead_time):
                        may_col = overlap_start_col - lead_time + d
                        if may_col < COL_DAY1:
                            continue
                        keikaku = may_ws.cell(may_block['計画（倍）'], may_col).value
                        if keikaku:
                            safe_write(jun_ws, jun_block['入庫予定数'], COL_DAY1 + d, keikaku * lot_size)
                            if b_idx == 0:
                                nyuko_written_dates.append(jun_start + timedelta(days=d))

            if nyuko_written_dates:
                first = nyuko_written_dates[0]
                last  = nyuko_written_dates[-1]
                if first == last:
                    label = f'入庫予定数（{first.month}/{first.day}）'
                else:
                    label = f'入庫予定数（{first.month}/{first.day}〜{last.month}/{last.day}）'
                item['transferred_types'].append(label)

            result['transferred'].append(item)

        jun_ws.protection.sheet = False

        out = BytesIO()
        jun_wb.save(out)
        out.seek(0)
        result['jun_bytes'] = out.getvalue()
        result['success'] = True

    except Exception as e:
        result['error'] = f'予期しないエラーが発生しました: {e}'
        result['detail'] = traceback.format_exc()

    return result


HTML = '''<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>月次引き継ぎ</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: "Hiragino Sans", "Yu Gothic", sans-serif; background: #f4f6f9; color: #222; }
  .container { max-width: 860px; margin: 40px auto; padding: 0 20px 60px; }
  h1 { font-size: 1.4rem; font-weight: 700; margin-bottom: 6px; }
  .sub { color: #666; font-size: 0.9rem; margin-bottom: 32px; }

  .card { background: white; border-radius: 10px; padding: 28px; margin-bottom: 20px; box-shadow: 0 1px 4px rgba(0,0,0,0.08); }
  .card h2 { font-size: 1rem; font-weight: 700; margin-bottom: 18px; color: #444; }

  .upload-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
  .upload-box label { display: block; font-size: 0.82rem; font-weight: 600; color: #555; margin-bottom: 6px; }
  .upload-box input[type=file] { width: 100%; border: 2px dashed #ccd; border-radius: 8px; padding: 12px; font-size: 0.85rem; cursor: pointer; background: #fafbff; }
  .upload-box input[type=file]:hover { border-color: #4472c4; }

  button[type=submit] { margin-top: 20px; width: 100%; padding: 14px; background: #4472c4; color: white; font-size: 1rem; font-weight: 700; border: none; border-radius: 8px; cursor: pointer; }
  button[type=submit]:hover { background: #3560b0; }

  .badge { display: inline-block; padding: 2px 9px; border-radius: 20px; font-size: 0.75rem; font-weight: 700; }
  .badge-ok   { background: #e6f4ea; color: #2d7a3a; }
  .badge-new  { background: #fff3cd; color: #856404; }
  .badge-warn { background: #fff3cd; color: #856404; }
  .badge-skip { background: #f0f0f0; color: #666; }
  .badge-err  { background: #fde8e8; color: #b91c1c; }

  .error-box { background: #fde8e8; border-left: 4px solid #ef4444; border-radius: 6px; padding: 16px 20px; margin-bottom: 20px; }
  .error-box strong { color: #b91c1c; }
  .error-box pre { margin-top: 10px; font-size: 0.75rem; background: #fff5f5; padding: 10px; border-radius: 4px; overflow-x: auto; white-space: pre-wrap; }

  .info-bar { background: #eef2fb; border-radius: 8px; padding: 12px 16px; font-size: 0.85rem; color: #3a3a6a; margin-bottom: 20px; display: flex; gap: 24px; flex-wrap: wrap; }
  .info-bar span strong { font-weight: 700; }

  .warnings-box { background: #fffbeb; border-left: 4px solid #f59e0b; border-radius: 6px; padding: 14px 18px; margin-bottom: 20px; }
  .warnings-box h3 { font-size: 0.88rem; font-weight: 700; color: #92400e; margin-bottom: 8px; }
  .warnings-box ul { padding-left: 18px; font-size: 0.84rem; color: #78350f; line-height: 1.8; }

  table { width: 100%; border-collapse: collapse; font-size: 0.85rem; }
  th { background: #f0f4fb; text-align: left; padding: 9px 12px; font-weight: 700; color: #555; border-bottom: 2px solid #dde3f0; }
  td { padding: 9px 12px; border-bottom: 1px solid #eee; vertical-align: top; }
  tr:last-child td { border-bottom: none; }
  .mono { font-family: monospace; font-size: 0.82rem; }

  .section-title { font-size: 0.82rem; font-weight: 700; color: #888; text-transform: uppercase; letter-spacing: 0.06em; margin: 28px 0 10px; }

  .download-btn { display: inline-block; margin-top: 20px; padding: 12px 28px; background: #16a34a; color: white; font-size: 0.95rem; font-weight: 700; border-radius: 8px; text-decoration: none; }
  .download-btn:hover { background: #15803d; }
</style>
</head>
<body>
<div class="container">
  <h1>月次引き継ぎスクリプト</h1>
  <p class="sub">今月のオーバーフロー計画（備考・計画（倍）・使用予測）と今月末在庫を来月ファイルに転記します。</p>

  <div class="card">
    <h2>ファイルを選択</h2>
    <form method="POST" enctype="multipart/form-data">
      <div class="upload-grid">
        <div class="upload-box">
          <label>今月のファイル（記入済み）（例: 6月計画.xlsx）</label>
          <input type="file" name="may_file" accept=".xlsx" required>
        </div>
        <div class="upload-box">
          <label>来月のファイル（空白）（例: 7月計画.xlsx）</label>
          <input type="file" name="jun_file" accept=".xlsx" required>
        </div>
      </div>
      <button type="submit">引き継ぎを実行</button>
    </form>
  </div>

  {% if result %}

    {% if result.error %}
    <div class="error-box">
      <strong>エラー: {{ result.error }}</strong>
      {% if result.detail %}<pre>{{ result.detail }}</pre>{% endif %}
    </div>

    {% else %}

    <div class="info-bar">
      <span><strong>今月:</strong> {{ result.may_start }} 〜 {{ result.may_last_date }}</span>
      <span><strong>来月:</strong> {{ result.jun_start }} 〜</span>
      <span><strong>転記済み:</strong> {{ result.transferred|length }} 商品</span>
      <span><strong>新規:</strong> {{ result.new_products|length }} 商品</span>
      <span><strong>廃止:</strong> {{ result.discontinued|length }} 商品</span>
    </div>

    {% if result.warnings %}
    <div class="warnings-box">
      <h3>⚠ 確認が必要な項目 ({{ result.warnings|length }}件)</h3>
      <ul>{% for w in result.warnings %}<li>{{ w }}</li>{% endfor %}</ul>
    </div>
    {% endif %}

    <a class="download-btn" href="/download">来月のファイルをダウンロード（記入済み）</a>

    {% if result.transferred %}
    <div class="section-title">転記済み商品</div>
    <div class="card" style="padding:0;overflow:hidden">
      <table>
        <thead><tr>
          <th>コード</th><th>商品名</th><th>棚卸し前在庫</th><th>転記した行</th><th>スキップ</th>
        </tr></thead>
        <tbody>
        {% for item in result.transferred %}
        <tr>
          <td class="mono">{{ item.code }}</td>
          <td>{{ item.name }}</td>
          <td>
            {% if item.takadoshi_mae is none %}
              <span class="badge badge-warn">空白</span>
            {% else %}
              {{ item.takadoshi_mae }}
              {% if item.takadoshi_warning %}<span class="badge badge-warn">要確認</span>{% endif %}
            {% endif %}
          </td>
          <td>
            {% for t in item.transferred_types %}
              <span class="badge badge-ok">{{ t }}</span>
            {% endfor %}
          </td>
          <td>
            {% for s in item.skipped_types %}
              <span class="badge badge-skip">{{ s }}</span>
            {% endfor %}
          </td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
    </div>
    {% endif %}

    {% if result.new_products %}
    <div class="section-title">新規商品（今月データなし・手動入力が必要）</div>
    <div class="card" style="padding:0;overflow:hidden">
      <table>
        <thead><tr><th>コード</th><th>商品名</th><th>対応</th></tr></thead>
        <tbody>
        {% for p in result.new_products %}
        <tr>
          <td class="mono">{{ p.code }}</td>
          <td>{{ p.name }}</td>
          <td><span class="badge badge-new">在庫数を手動入力してください</span></td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
    </div>
    {% endif %}

    {% if result.discontinued %}
    <div class="section-title">廃止商品（来月ファイルに存在しないためスキップ）</div>
    <div class="card" style="padding:0;overflow:hidden">
      <table>
        <thead><tr><th>コード</th><th>商品名</th></tr></thead>
        <tbody>
        {% for p in result.discontinued %}
        <tr>
          <td class="mono">{{ p.code }}</td>
          <td>{{ p.name }}</td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
    </div>
    {% endif %}

    {% endif %}
  {% endif %}
</div>
</body>
</html>'''


_last_result_bytes = None
_last_jun_filename = 'jun_output移行済.xlsx'


@app.route('/', methods=['GET', 'POST'])
def index():
    global _last_result_bytes, _last_jun_filename
    result = None

    if request.method == 'POST':
        may_file = request.files.get('may_file')
        jun_file = request.files.get('jun_file')

        if not may_file or not jun_file:
            result = {'error': 'ファイルが選択されていません。'}
        else:
            result = run_handoff(may_file.read(), jun_file.read())
            if result.get('jun_bytes'):
                _last_result_bytes = result['jun_bytes']
                stem = (jun_file.filename or 'jun_output').rsplit('.', 1)[0]
                _last_jun_filename = f'{stem}移行済.xlsx'

    return render_template_string(HTML, result=result)


@app.route('/download')
def download():
    if not _last_result_bytes:
        return 'ファイルがありません。先に引き継ぎを実行してください。', 404
    return send_file(
        BytesIO(_last_result_bytes),
        as_attachment=True,
        download_name=_last_jun_filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    print("起動中: http://localhost:5001")
    app.run(debug=True, port=5001)
